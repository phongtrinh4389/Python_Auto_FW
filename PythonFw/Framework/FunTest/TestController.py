__author__ = 'phongtrinh'

import nose
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import Style, PatternFill
import os
import inspect
import shutil
import time
import datetime


def fix_path(src_path, destination_path):
    if os.path.dirname(destination_path) == '':
        return os.path.join(os.path.dirname(src_path), destination_path)
    else:
        return destination_path


def cp(src_path, destination_path, *args):
    """Copy a file or tree"""
    destination_path = fix_path(src_path, destination_path)
    if 'r' in args:
        shutil.copytree(src_path, destination_path)
    else:
        shutil.copy(src_path, destination_path)


def get_path():
    this_path = os.path.split(inspect.getfile(inspect.currentframe()))[0]
    return os.path.realpath(os.path.abspath(this_path))


def get_project_root():
    current = get_path()
    root = os.path.dirname(os.path.dirname(current))
    return root


def get_file_path(*args):
    """Take location of the projecy root and add subdirectories + filename"""
    root = get_project_root()
    return os.path.join(root, 'Automated Tests', *args)


def create_directory(path):
    if os.path.exists(path):
        return
    os.mkdir(path)


def write_excel(filename, sheet_name, column, row, data, hyperlink=None, image=None):
    try:
        writer = openpyxl.load_workbook(filename)
        sheet = writer.get_sheet_by_name(sheet_name)
        sheet.cell(column + str(row)).value = data
        if data == 'PASS':
            fill = PatternFill(patternType='solid', fgColor=colors.GREEN)
            sheet.cell(column + str(row)).style = Style(fill=fill)
        elif data == 'FAIL':
            fill = PatternFill(patternType='solid', fgColor=colors.RED)
            sheet.cell(column + str(row)).style = Style(fill=fill)
        else:
            pass
        if hyperlink is not None:
            sheet.cell(column + str(row)).value = '=HYPERLINK("%s","%s")' % (hyperlink, data)
            font = Font(italic=True, underline=Font.UNDERLINE_SINGLE, color=colors.BLUE)
            sheet.cell(column + str(row)).font = font
        writer.save(filename)
    except Exception:
        pass


class TestController(object):
    # The root test folder
    root_tests = 'Automated Tests'
    # The common report folder
    report_folder_name = 'Report'
    # The name of image report folder
    image_report_folder_name = 'image_report'
    # The name of html report folder
    html_report_folder_name = 'html_report'
    # The name of excel report folder
    excel_report_folder_name = 'excel_report'
    # PASS/FAIL constants
    PASS = 'PASS'
    FAIL = 'FAIL'
    # BLOCK/DONE constants
    BLOCK = 'Block'
    NA = 'NA'
    DONE = 'Done'
    # The script name column in the Test_Cases.xlsx
    script_name_column = 'G'
    # The status of test scripts column in the Test_Cases.xlsx (Blocked/Done)
    status_column = 'H'
    # The html report column in the Test_Cases.xlsx
    html_column = 'J'
    # The image report column in the Test_Cases.xlsx
    image_column = 'L'
    # The result column in the Test_Cases.xlsx
    result_column = 'I'
    # The running time column in the Test_Cases.xlsx
    run_time_column = 'K'
    # Date time format
    date_format = '%d_%m_%y'
    # The template name of image report
    image_template = 'test.jpeg'

    def __init__(self, test_repo, date_time=''):
        # The name of test repository
        self.test_repo = test_repo
        if date_time is '':
            self.date_time = time.strftime(self.date_format)
        else:
            try:
                datetime.datetime.strptime(date_time, self.date_format)
                self.date_time = date_time
            except ValueError as e:
                raise e
        self.report_path = self.get_report_path(self.test_repo, self.date_time)

    def get_report_path(self, repo, date_time):
        report_path = get_file_path(repo, self.report_folder_name,
                                    '%s_Report_%s' % (self.test_repo, date_time))
        create_directory(report_path)
        return report_path

    def get_excel_report_path(self):
        report_path = self.report_path
        excel_report_path = report_path + '\\' + self.excel_report_folder_name
        create_directory(excel_report_path)
        return excel_report_path + '\\'

    def get_html_report_path(self):
        report_path = self.report_path
        html_report_path = report_path + '\\' + self.html_report_folder_name
        create_directory(html_report_path)
        return html_report_path + '\\'

    def get_image_report_path(self):
        report_path = self.report_path
        image_report_path = report_path + '\\' + self.image_report_folder_name
        create_directory(image_report_path)
        return image_report_path + '\\'

    def move_image_report_file(self, destination_file, source_file=image_template):
        image_path = self.get_image_report_path()
        image_src = get_file_path(get_project_root(), self.root_tests, source_file)
        try:
            shutil.move(image_src, image_path + destination_file + '.jpeg')
            if os.path.isfile(image_src):
                os.remove(image_src)
        except FileNotFoundError as e:
            pass

    def run(self, args):
        begin_time = time.mktime(datetime.datetime.now().timetuple())
        result = nose.run(argv=args)
        end_time = time.mktime(datetime.datetime.now().timetuple())
        duration_time = end_time - begin_time
        if result:
            return [self.PASS, duration_time]
        else:
            return [self.FAIL, duration_time]

    def run_test_suite(self, filename, suite_name, re_run=False):
        html_folder = self.html_report_folder_name
        image_folder = self.image_report_folder_name
        # Generate excel report into excel report folder
        excel_report_path = self.get_excel_report_path()
        excel_report_file_path = '%s/%s_Result_%s.xlsx' % (excel_report_path, self.test_repo, self.date_time)
        if os.path.isfile(excel_report_file_path) is False:
            cp(filename, excel_report_file_path)

        reader = openpyxl.load_workbook(excel_report_file_path)
        sheet = reader.get_sheet_by_name(suite_name)
        row_index = 1
        for row in sheet.iter_rows():
            script_status = sheet.cell(self.status_column + '%s' % row_index).value
            script_name = sheet.cell(self.script_name_column + '%s' % row_index).value
            if row_index == 1:
                row_index += 1
                continue
            elif script_status == self.BLOCK:
                write_excel(excel_report_file_path, suite_name, self.result_column, row_index, self.BLOCK)
                row_index += 1
                continue
            elif script_status == self.NA:
                write_excel(excel_report_file_path, suite_name, self.result_column, row_index, self.NA)
                row_index += 1
                continue
            elif script_name is None:
                row_index += 1
                continue
            elif script_name.endswith('_Tests') is False:
                row_index += 1
                continue
            else:
                if re_run is True and sheet.cell(self.result_column + '%s' % row_index).value == self.PASS:
                    row_index += 1
                    continue
                html_report_path = self.get_html_report_path()
                args = ['-v', '--with-html', '--html-file=%s/%s.html' % (html_report_path, script_name),
                        'Plan/%s.py' % script_name]
                # args = ['-v', 'Plan/%s.py'%script_name]
                result = self.run(args)
                self.move_image_report_file(script_name)
                write_excel(excel_report_file_path, suite_name, self.html_column, row_index, '%s.html' % script_name,
                            hyperlink='..\%s\%s.html' % (html_folder, script_name))
                write_excel(excel_report_file_path, suite_name, self.image_column, row_index, '%s.jpeg' % script_name,
                            hyperlink='..\%s\%s.jpeg' % (image_folder, script_name))
                write_excel(excel_report_file_path, suite_name, self.result_column, row_index, result[0])
                write_excel(excel_report_file_path, suite_name, self.run_time_column, row_index, result[1])
                row_index += 1


